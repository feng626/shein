# -*- coding: utf-8 -*-
#
from io import BytesIO

from openpyxl.reader.excel import load_workbook
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.response import Response

from orgs.mixins.api import OrgBulkModelViewSet
from ..models import Product
from ..serializers import ProductSerializer, ImportExcelSerializer

__all__ = ['ProductViewSet']


class ProductViewSet(OrgBulkModelViewSet):
    model = Product
    filterset_fields = ("name", 'spu', 'skc', 'sku', 'type', 'cost', 'price')
    search_fields = ('type',)
    serializer_classes = {
        'default': ProductSerializer,
        'sync': ImportExcelSerializer,
    }

    rbac_perms = {
        'sync': 'products.change_product',
    }

    @action(methods=['patch'], detail=False, url_path='sync')
    def sync(self, request, *args, **kwargs):
        serializer = ImportExcelSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        upload_file = serializer.validated_data['file']

        workbook = load_workbook(filename=BytesIO(upload_file.read()), data_only=True)
        sheet = workbook.active
        # ('l250211172600', '定制积木&拼图玩具', '否', 'sl25021117260016136', 'I92s7amrik14', '颜色', '1', '尺寸', '均码',
        #  '', '', '', '', '',
        #  'Custom Birthday Gift for Him – Personalized Brick Figure with Frame | Custom LEGO-Style Mini Block Model | Small Particle Building Toy',
        #  'SDLasdhf', '', '',
        #  'https://img.ltwebstatic.com/images3_spmp/2025/02/11/15/17392461879e17510bdc2283f8c5e387a9aa58dbe2_square.jpg',
        #  '', '', '', '', '', '', '', '', '', '', '', '', '', '99', '250', 'g', '18.00', '18.00', '7.00', 'cm', '150.00')

        for row in sheet.iter_rows(min_row=2, values_only=True):  # 跳过表头
            if len(row) != 40:
                return Response(data={"msg": "Invalid file"}, status=status.HTTP_400_BAD_REQUEST)

            (
                spu, tp, __, skc, sku, __, __, __, __, __,
                __, __, __, __, name, number, __, __, avatar_url
            ) = row[:19]
            price = row[-1]

            self.model.objects.update_or_create(
                sku=sku,
                defaults={
                    'name': name,
                    'spu': spu,
                    'type': tp,
                    'skc': skc,
                    'number': number,
                    'avatar_url': avatar_url,
                    'price': price,
                }
            )

        return Response(data={"msg": "Success"}, status=status.HTTP_200_OK)
