# -*- coding: utf-8 -*-
#
from datetime import datetime as dt
from io import BytesIO

from django.utils import timezone
from openpyxl.reader.excel import load_workbook
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.response import Response

from common.drf.filters import DatetimeRangeFilterBackend
from orgs.mixins.api import OrgBulkModelViewSet
from products.models import Product
from products.serializers import ImportExcelSerializer
from ..models import Transaction
from ..serializers import TransactionListSerializer, TransactionSerializer

__all__ = ['TransactionViewSet']


class TransactionViewSet(OrgBulkModelViewSet):
    model = Transaction
    filterset_fields = ("name", "amount", "price")
    search_fields = ("name",)
    extra_filter_backends = [DatetimeRangeFilterBackend]
    date_range_filter_fields = [
        ('datetime', ('date_from', 'date_to'))
    ]
    serializer_classes = {
        'default': TransactionSerializer,
        'list': TransactionListSerializer,
        'sync': ImportExcelSerializer,
    }
    rbac_perms = {
        'list': 'finances.view_transaction',
        'sync': 'finances.change_transaction',
        'profit': 'finances.view_transaction',
    }

    @action(methods=['get'], detail=False, url_path='profit')
    def profit(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        profit = 0
        product_cost_map = {str(_id): cost for _id, cost in Product.objects.values_list('id', 'cost')}

        for product_id, amount, price in queryset.values_list('product_id', 'amount', 'price'):
            profit += (price - product_cost_map[str(product_id)]) * amount

        return Response(data={"msg": "Success", "profit": str(profit)}, status=status.HTTP_200_OK)

    @action(methods=['patch'], detail=False, url_path='sync')
    def sync(self, request, *args, **kwargs):
        serializer = ImportExcelSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        upload_file = serializer.validated_data['file']

        workbook = load_workbook(filename=BytesIO(upload_file.read()), data_only=True)
        sheet = workbook.active
        # ('B2503020270129158', '9789*g-Multicolor', 'sw2412160037244105', 'I61fzx8xtsb4', 'sw2412160037244105-007',
        #  '多色-Samsung Galaxy Z Flip4', 1, '平台客单发货', '收入', '4334116337559562', 'GSH1MX11100M8LR', '2025-03-02', '2025-03-02',
        #  '2025-03-30', '待结算', '/', 'CNY', 4.8, 'INFINITE STYLES SERVICES CO., LIMITED',
        #  '客单创建时间:2025-02-26 04:19:40,\n参与活动:2025年SHEIN春促活动-跨境,\n活动时间范围:2025-02-17 00:00:00~2025-03-10 23:59:59,\n活动调价方式:定价,\n提报的活动价格:4.80,\n结算价格:4.80')
        product_sku_map = {sku: str(_id) for _id, sku in Product.objects.values_list('id', 'sku')}
        not_exist_list = []
        for row in sheet.iter_rows(min_row=2, values_only=True):  # 跳过表头
            if len(row) != 20:
                return Response(data={"msg": "Invalid file"}, status=status.HTTP_400_BAD_REQUEST)

            (
                name, number, skc, sku, merchants_sku, attr, amount, __, __,
                business_number, source_number, datetime_str, ledger_time,
                settlement_time, state, actual_settlement_time, __, price, company, other
            ) = row

            comment = "\n".join([
                f"货号: {number}",
                f"SKC: {skc}",
                f"商家SKU: {merchants_sku}",
                f"属性集: {attr}",
                f"来源单号: {source_number}",
                f"台账添加时间: {ledger_time}",
                f"预计结算日期: {settlement_time}",
                f"实际结算日期: {actual_settlement_time}",
                f"公司主体: {company}",
                f"活动信息: {other}"
            ])

            if sku not in product_sku_map:
                not_exist_list.append({
                    'sku': sku,
                    'business_number': business_number,
                })
                continue

            product_id = product_sku_map[sku]

            date_value = timezone.make_aware(dt.strptime(datetime_str, "%Y-%m-%d"))
            self.model.objects.update_or_create(
                business_number=business_number,
                defaults={
                    'name': name,
                    'amount': amount,
                    'price': price,
                    'state': state,
                    'source_number': source_number,
                    'product_id': product_id,
                    'datetime': date_value,
                    'comment': comment,
                }
            )

        return Response(data={"msg": "Success", "not_exists": not_exist_list}, status=status.HTTP_200_OK)
